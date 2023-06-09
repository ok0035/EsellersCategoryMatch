import os
# 엑셀 파일 읽기
import xlrd
import openpyxl
# 엑셀 파일 생성
from xlutils.copy import copy
from xlwt import Workbook

# Directory containing the Excel files
coupang_path = '쿠팡'
download_set_path = '세트'


def main():
    # market_category_excel = load_excel("마켓카테고리매칭정보.xls")

    download_set = load_excel_xls(os.path.join(download_set_path, "ZSM_20230316_esellers.xls"))
    download_set_default_sheet = download_set.sheet_by_name('기본정보')

    wt_download_set: Workbook = copy(download_set)
    extension_sheet = wt_download_set.get_sheet(1)

    # 마켓 카테고리 시트 (원하는 마켓으로 시트 선택)
    # market_category_sheet = market_category_excel.sheet_by_name("쿠팡")

    # 다운로드 세트 확장정보
    # 이셀러스 카테고리 번호와 매칭되는 마켓 카테고리 번호를 여기에 저장

    for row in range(2, download_set_default_sheet.nrows):
        set_category_num: str = download_set_default_sheet.cell(row, 3).value
        set_category_name = download_set_default_sheet.cell(row, 4).value
        esellers_categories = findEsellersCategory(set_category_num)
        esellers_category_name: str = ""

        for category in esellers_categories:
            esellers_category_name.join(category + ">")

        coupang_category_num = getCoupangCategoryNumber(esellers_categories)
        extension_sheet.write(row, 7, coupang_category_num)
        # print(set_category_num, "\t", set_category_name, "\t", esellers_category_name, " -> ", coupang_category_num)

    wt_download_set.save("extension_category_set.xls")


# 이셀러스 카테고리 리스트 가져오기
def findEsellersCategory(set_category_number: str) -> list:
    category_excel_file = load_excel_xls("원본상품등록양식Ver.1.0.4.1.xls")
    esellers_category_sheet = category_excel_file.sheet_by_name("이셀러스표준카테고리")
    for row in range(esellers_category_sheet.nrows):
        if set_category_number == esellers_category_sheet.cell(row, 0).value:
            big_category: str = esellers_category_sheet.cell(row, 1).value
            medium_category: str = esellers_category_sheet.cell(row, 2).value
            small_category: str = esellers_category_sheet.cell(row, 3).value
            detail_category: str = esellers_category_sheet.cell(row, 4).value
            # print("이셀러스 카테고리 ", set_category_number, big_category, medium_category, small_category, detail_category)
            return [detail_category, small_category, medium_category, big_category]

    return []


# 쿠팡 카테고리 번호
def getCoupangCategoryNumber(esellers_cat_list: list) -> str:
    path = os.path.join(coupang_path, 'union_coupang_category.xls')
    if check_file_exists(path) is False:
        print("쿠팡 통합 카테고리 엑셀 파일이 없습니다.")
        create_coupang_category_file()

    coupang_categories_file = load_excel_xls(os.path.join(coupang_path, 'union_coupang_category.xls'))
    coupang_categories_sheet = coupang_categories_file.sheet_by_index(0)

    # print("이셀러스 카테고리 리스트 -> ", esellers_cat_list, coupang_categories_sheet.nrows)
    esellers_cat_list.reverse()
    return searchCategoryByCoupang(esellers_cat_list, coupang_categories_sheet)


def searchCategoryByCoupang(esellers_cat_list: list, coupang_sheet: xlrd.sheet.Sheet) -> str:
    assert isinstance(coupang_sheet, xlrd.sheet.Sheet)

    print("================================================================================================================================")
    print("이셀러스 카테고리 ", esellers_cat_list, "에 맞는 쿠팡 카테고리 찾는 중...")

    max_count = 0
    match_category_num = ""
    match_category = ""

    for row in range(coupang_sheet.nrows):
        coupang_category_cell: str = coupang_sheet.cell(row, 1).value
        coupang_categories = coupang_category_cell.split(">")

        match_count = 0

        for c_category_index in range(len(coupang_categories)):
            c_categories = coupang_categories[c_category_index]
            c_split_categories = c_categories.split("/")
            for c_category in c_split_categories:

                assert isinstance(c_category, str)

                for e_category_index in range(len(esellers_cat_list)):
                    e_categories = esellers_cat_list[e_category_index]
                    assert isinstance(e_categories, str)
                    bonus_point = pow(abs(e_category_index + c_category_index), 2)
                    categories = e_categories.split("/")

                    for e_category_split_index in range(0, len(categories)):
                        e_category = categories[e_category_split_index]
                        if e_category == "":
                            continue

                        if e_category == c_categories:
                            match_count += bonus_point + 40
                        elif e_category == c_category:
                            match_count += bonus_point + 20
                        elif e_category in c_category:
                            match_count += bonus_point + 3
                        else:
                            match_count -= 1

        if match_count > max_count:
            print(match_count, "포인트를 획득하여", match_category, "에서 ", coupang_sheet.cell(row, 1).value, "로 변경되었습니다.")
            match_category_num = coupang_sheet.cell(row, 0).value
            match_category = coupang_sheet.cell(row, 1).value
            max_count = match_count

    print("최종으로 매칭된 카테고리 : ", esellers_cat_list,  " -> ", match_category, match_category_num)
    print("================================================================================================================================")

    return match_category_num


    # 이셀러스 카테고리 리스트를 소분류부터 사용하기 위해 역순으로 인덱스를 가져옴
    # for e_index in range(len(esellers_cat_list) - 1, 0, -1):
    #     e_category: str = esellers_cat_list[e_index]
    #     e_categories: list = e_category.split("/")
    #     is_exist = True  # 검색할 카테고리가 존재하는지 확인
    #     for category in e_categories:
    #         i = 1  # 쿠팡 카테고리와 이셀러스 카테고리가 일치하는지 소분류부터 검색
    #         while True:
    #
    #             for row in range(coupang_sheet.nrows):
    #                 coupang_category_cell: str = coupang_sheet.cell(row, 1).value
    #                 coupang_categories = coupang_category_cell.split(">")
    #                 coupang_category_index = len(coupang_categories) - i
    #
    #                 is_exist = False
    #                 if coupang_category_index < 0:
    #                     continue
    #
    #                 is_exist = True
    #                 coupang_ca: str = coupang_categories[coupang_category_index]
    #
    #                 # 소 분류 카테고리가 /로 나누어져 있을 경우 하나씩 검색
    #                 split_categories = coupang_ca.split("/")
    #
    #                 for coupang_category in split_categories:
    #                     if category in coupang_category:
    #                         print(esellers_cat_list, "\t\t -> \t\t", coupang_category_cell)
    #                         return coupang_sheet.cell(row, 0).value
    #
    #             if not is_exist:
    #                 break
    #
    #             i = i + 1
    # return ""


# 쿠팡 통합 카테고리 파일 생성
def create_coupang_category_file():
    print("쿠팡 통합 카테고리 파일을 생성중입니다...")
    # Create a new workbook

    # 쿠팡 엑셀 파일 읽기
    # Loop through all the Excel files in the directory
    row = 0
    # Create a new worksheet in the new workbook
    union_coupang_category_file = xlwt.Workbook()
    coupang_category_sheet = union_coupang_category_file.add_sheet("category")
    for file_name in os.listdir(coupang_path):
        print(file_name, "파일 통합 중...")
        if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
            # Open the Excel file
            workbook = load_excel_xlsx(os.path.join(coupang_path, file_name))
            worksheet = workbook['data']

            for cell in worksheet['A']:
                value: str = cell.value
                if value[0] == '[':
                    categories = value.split(' ')
                    if len(categories) >= 2:
                        number = categories[0].replace('[', '').replace(']', '')
                        category = categories[1]
                        coupang_category_sheet.write(row, 0, number)
                        coupang_category_sheet.write(row, 1, category)
                        row += 1

    union_coupang_category_file.save(os.path.join(coupang_path, 'union_coupang_category.xls'))


# 통합 카테고리 파일이 존재하는지 확인
def check_file_exists(file_path):
    if os.path.exists(file_path):
        try:
            xlrd.open_workbook(file_path)
            return True
        except xlrd.XLRDError:
            return False
    else:
        return False


# xls 파일 열기
def load_excel_xls(file):
    return xlrd.open_workbook(filename=file)


# xlsx 파일 열기
def load_excel_xlsx(file):
    return openpyxl.load_workbook(filename=file)


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    main()
